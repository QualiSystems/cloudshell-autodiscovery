import string

import xlsxwriter
from openpyxl import load_workbook

from autodiscovery.reports.base import AbstractParsableReport


class AbstractExcelReport(AbstractParsableReport):
    FILE_EXTENSION = "xlsx"

    def _prepare_column(self, start_column, end_column=None):
        """Prepare column name.

        :param str start_column:
        :param str end_column:
        :return:
        """
        if end_column is None:
            end_column = start_column

        return "{}:{}".format(start_column, end_column)

    @property
    def _header_with_column(self):
        return zip(self._header, string.ascii_uppercase)

    @property
    def _header_column_width_map(self):
        return {}

    def _format_columns_width(self, worksheet):
        """Format columns width.

        :param worksheet:
        :return:
        """
        for header, column in self._header_with_column:
            if header in self._header_column_width_map:
                worksheet.set_column(
                    self._prepare_column(column), self._header_column_width_map[header]
                )

    def _get_cell_value(self, wb_sheet, column, row):
        """Get cell value.

        :param wb_sheet:
        :param column:
        :param row:
        :return:
        """
        cell = wb_sheet["{}{}".format(column, row)]
        return cell.value or ""

    def parse_entries_from_file(self, report_file):
        """Parse entries from file.

        :param str report_file: path to the report file
        :rtype: list[Entry]
        """
        wb = load_workbook(report_file)
        wb_sheet = wb.active

        for row_num in xrange(2, wb_sheet.max_row + 1):  # first row is a header  # noqa
            entry_attrs = {}
            for header, column in self._header_with_column:
                entry_attr = self._header_entry_map[header]
                entry_attrs[entry_attr] = self._get_cell_value(
                    wb_sheet, column, row_num
                )

            entry = self.entry_class(**entry_attrs)
            self._entries.append(entry)

    def generate(self):
        """Save report for all discovered devices into the excel file."""
        # todo(A.Piddubny): use one library to read/write xlsx files - openpyxl
        workbook = xlsxwriter.Workbook(self.file_name)
        worksheet = workbook.add_worksheet()
        table_data = [self._header]

        for entry in self._entries:
            entry_row = [
                getattr(entry, attr) for attr in self._header_entry_map.values()
            ]
            table_data.append(entry_row)

        for row_num, row in enumerate(table_data):
            for col_num, col in enumerate(row):
                worksheet.write(row_num, col_num, col)

        # set bold header
        worksheet.set_row(0, None, workbook.add_format({"bold": True}))

        # format columns width
        self._format_columns_width(worksheet)

        workbook.close()
