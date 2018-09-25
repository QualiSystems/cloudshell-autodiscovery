import xlsxwriter
from openpyxl import load_workbook

from autodiscovery.reports.connections.base import AbstractReport
from autodiscovery.reports.connections.base import Entry


class ExcelReport(AbstractReport):
    FILE_EXTENSION = ".xlsx"
    DEFAULT_REPORT_FILE = "connect_ports_report{}".format(FILE_EXTENSION)

    SOURCE_PORT_COLUMN = "A"
    TARGET_PORT_COLUMN = "B"
    DOMAIN_COLUMN = "C"
    STATUS_COLUMN = "D"
    COMMENT_COLUMN = "E"

    def __init__(self, file_name=None):
        """

        :param str file_name:
        """
        super(ExcelReport, self).__init__()

        # todo: this code is duplicated
        if file_name is None:
            file_name = self.DEFAULT_REPORT_FILE
        elif not file_name.lower().endswith(self.FILE_EXTENSION):
            file_name += self.FILE_EXTENSION

        self.file_name = file_name

    def generate(self):
        """Save report for all discovered devices into the excel file"""
        # todo(A.Piddubny): use one library to read/write xlsx files - openpyxl
        workbook = xlsxwriter.Workbook(self.file_name)
        worksheet = workbook.add_worksheet()
        table_data = [self.HEADER]

        for entry in self._entries:
            table_data.append((entry.source_port, entry.target_port, entry.domain, entry.status, entry.comment))

        for row_num, row in enumerate(table_data):
            for col_num, col in enumerate(row):
                worksheet.write(row_num, col_num, col)

        # set bold header
        worksheet.set_row(0, None, workbook.add_format({'bold': True}))

        def prepare_column(start_column, end_column=None):
            """

            :param str start_column:
            :param str end_column:
            :return:
            """
            if end_column is None:
                end_column = start_column

            return "{}:{}".format(start_column, end_column)

        # format columns width
        worksheet.set_column(prepare_column(self.SOURCE_PORT_COLUMN), 30)
        worksheet.set_column(prepare_column(self.TARGET_PORT_COLUMN), 30)
        worksheet.set_column(prepare_column(self.DOMAIN_COLUMN), 20)
        worksheet.set_column(prepare_column(self.STATUS_COLUMN), 25)
        worksheet.set_column(prepare_column(self.COMMENT_COLUMN), 40)

        workbook.close()

    def parse_entries_from_file(self, report_file):
        """

        :param str report_file: path to the report file
        :rtype: list[Entry]
        """
        entries = []
        wb = load_workbook(report_file)
        wb_sheet = wb.active

        def get_cell_value(column, row):
            cell = wb_sheet["{}{}".format(column, row)]
            return cell.value or ""

        for row_num in xrange(2, wb_sheet.max_row+1):  # first row is a header
            entry = Entry(source_port=get_cell_value(ExcelReport.SOURCE_PORT_COLUMN, row_num),
                          target_port=get_cell_value(ExcelReport.TARGET_PORT_COLUMN, row_num),
                          domain=get_cell_value(ExcelReport.DOMAIN_COLUMN, row_num),
                          status=get_cell_value(ExcelReport.STATUS_COLUMN, row_num),
                          comment=get_cell_value(ExcelReport.COMMENT_COLUMN, row_num))

            entries.append(entry)

        return entries
