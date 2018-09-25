import re

import xlsxwriter
from openpyxl import load_workbook

from autodiscovery.reports.discovery.base import AbstractReport
from autodiscovery.reports.discovery.base import Entry


class ExcelReport(AbstractReport):
    FILE_EXTENSION = ".xlsx"
    DEFAULT_REPORT_FILE = "discovery-report{}".format(FILE_EXTENSION)

    IP_COLUMN = "A"
    VENDOR_COLUMN = "B"
    SYS_OBJ_COLUMN = "C"
    DESCRIPTION_COLUMN = "D"
    SNMP_COMMUNITY_COLUMN = "E"
    MODEL_TYPE_COLUMN = "F"
    DEVICE_NAME_COLUMN = "G"
    DOMAIN_COLUMN = "H"
    FOLDER_COLUMN = "I"
    ATTRIBUTES_COLUMN = "J"
    STATUS_COLUMN = "K"
    COMMENT_COLUMN = "L"

    def __init__(self, file_name=None):
        """

        :param str file_name:
        """
        super(ExcelReport, self).__init__()

        if file_name is None:
            file_name = self.DEFAULT_REPORT_FILE
        elif not file_name.lower().endswith(self.FILE_EXTENSION):
            file_name += self.FILE_EXTENSION

        self.file_name = file_name

    def generate(self):
        """Save report for all discovered devices into the excel file"""
        # todo(A.Piddubny): use one library to read/write xlsx files - openpyxl
        workbook = xlsxwriter.Workbook(self.file_name)
        worksheet = workbook.add_worksheet()
        table_data = [self.HEADER]

        for entry in self._entries:
            description = re.sub("\s+", " ", entry.description)  # replace all \n \r \t symbols
            table_data.append((entry.ip, entry.vendor, entry.sys_object_id, description, entry.snmp_community,
                               entry.model_type, entry.device_name, entry.domain, entry.folder_path,
                               entry.formatted_attrs, entry.status, entry.comment))

        for row_num, row in enumerate(table_data):
            for col_num, col in enumerate(row):
                worksheet.write(row_num, col_num, col)

        # set bold header
        worksheet.set_row(0, None, workbook.add_format({'bold': True}))

        def prepare_column(start_column, end_column=None):
            """

            :param str start_column:
            :param str end_column:
            :return:
            """
            if end_column is None:
                end_column = start_column

            return "{}:{}".format(start_column, end_column)

        # format columns width
        worksheet.set_column(prepare_column(self.IP_COLUMN, self.VENDOR_COLUMN), 20)
        worksheet.set_column(prepare_column(self.SYS_OBJ_COLUMN), 30)
        worksheet.set_column(prepare_column(self.DESCRIPTION_COLUMN), 50)
        worksheet.set_column(prepare_column(self.SNMP_COMMUNITY_COLUMN), 30)
        worksheet.set_column(prepare_column(self.MODEL_TYPE_COLUMN, self.DEVICE_NAME_COLUMN), 20)
        worksheet.set_column(prepare_column(self.DOMAIN_COLUMN), 20)
        worksheet.set_column(prepare_column(self.FOLDER_COLUMN), 20)
        worksheet.set_column(prepare_column(self.ATTRIBUTES_COLUMN), 25)
        worksheet.set_column(prepare_column(self.STATUS_COLUMN), 25)
        worksheet.set_column(prepare_column(self.COMMENT_COLUMN), 40)

        workbook.close()

    def parse_entries_from_file(self, report_file):
        """

        :param str report_file: path to the report file
        :rtype: list[Entry]
        """
        entries = []
        wb = load_workbook(report_file)
        wb_sheet = wb.active

        def get_cell_value(column, row):
            cell = wb_sheet["{}{}".format(column, row)]
            return cell.value or ""

        for row_num in xrange(2, wb_sheet.max_row+1):  # first row is a header
            formatted_attributes = get_cell_value(ExcelReport.ATTRIBUTES_COLUMN, row_num)

            entry = Entry(ip=get_cell_value(ExcelReport.IP_COLUMN, row_num),
                          vendor=get_cell_value(ExcelReport.VENDOR_COLUMN, row_num),
                          sys_object_id=get_cell_value(ExcelReport.SYS_OBJ_COLUMN, row_num),
                          description=get_cell_value(ExcelReport.DESCRIPTION_COLUMN, row_num),
                          snmp_community=get_cell_value(ExcelReport.SNMP_COMMUNITY_COLUMN, row_num),
                          model_type=get_cell_value(ExcelReport.MODEL_TYPE_COLUMN, row_num),
                          device_name=get_cell_value(ExcelReport.DEVICE_NAME_COLUMN, row_num),
                          domain=get_cell_value(ExcelReport.DOMAIN_COLUMN, row_num),
                          folder_path=get_cell_value(ExcelReport.FOLDER_COLUMN, row_num),
                          attributes=Entry.parse_formatted_attrs(formatted_attributes),
                          status=get_cell_value(ExcelReport.STATUS_COLUMN, row_num),
                          comment=get_cell_value(ExcelReport.COMMENT_COLUMN, row_num))

            entries.append(entry)

        return entries
